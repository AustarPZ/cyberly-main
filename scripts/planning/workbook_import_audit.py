import argparse
import csv
import hashlib
import json
import sys
from collections import Counter
from copy import deepcopy
from datetime import date, datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from scripts.planning import workbook_import_preview as preview


AUDIT_SCHEMA_VERSION = "1.0"
DECLARED_LOCALES = ["en", "ms", "zh-CN"]
AUDIT_OUTPUT_SUFFIX = "14July2026-audit"
REQUIRED_LIVE_MODELS = {"resource", "scenario"}
MISSING_CONTENT_CODES = {
    "missing_body",
    "missing_faq_answer",
    "missing_scenario_steps",
    "missing_scenario_options",
    "missing_safety_guidance",
}


def sha256_file(path):
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


def file_identity(path):
    path = Path(path)
    stat = path.stat()
    return {
        "path": path.as_posix(),
        "fileName": path.name,
        "sha256": sha256_file(path),
        "sizeBytes": stat.st_size,
        "lastModified": datetime.fromtimestamp(stat.st_mtime).isoformat(),
    }


def is_audit_output_path(path):
    return Path(path).name == AUDIT_OUTPUT_SUFFIX


def sanitize_audit_payload(value):
    return preview.sanitize_for_json(value)


def assert_catalogue_read_only(catalog):
    database = catalog.get("database", {})
    if database.get("countsBefore") != database.get("countsAfter") or database.get("unchanged") is not True:
        raise RuntimeError("Read-only catalogue counts changed.")


def content_issue_codes(item):
    return {issue.get("issue_code") for issue in item.get("validation", {}).get("issues", [])}


def classify_content_completeness(item):
    content_type = item.get("normalized", {}).get("contentType")
    normalized = item.get("normalized", {})
    title = bool(normalized.get("title"))
    summary = bool(normalized.get("summary"))
    body = bool(normalized.get("body"))
    issues = content_issue_codes(item)
    reasons = []

    if content_type == "unsupported_content_type" or not title:
        return "malformed", ["Record cannot be interpreted as a supported planning content type."]

    if content_type == "resource":
        if title and summary and body:
            return "full_content", ["Resource has title, summary, and learner-facing body content."]
        if title and summary:
            reasons.append("Resource has title/summary/source planning fields but no learner-facing body column.")
            return "content_brief", reasons
        return "metadata_only", ["Resource row lacks meaningful learner-facing content beyond metadata."]

    if content_type == "scenario":
        structure = item.get("scenarioStructure", {})
        if title and summary and structure.get("steps") and structure.get("options"):
            return "full_content", ["Scenario has title, description, steps, and options."]
        if title and summary:
            reasons.append("Scenario has title/description planning fields but no step/option structure.")
            if "missing_scenario_steps" in issues:
                reasons.append("Scenario steps are absent.")
            if "missing_scenario_options" in issues:
                reasons.append("Scenario options are absent.")
            return "content_brief", reasons
        return "metadata_only", ["Scenario row lacks learner-facing setup content."]

    if content_type == "faq":
        if title and body:
            return "full_content", ["FAQ has question/title and answer text."]
        if title or summary:
            return "content_brief", ["FAQ row has question/summary prompt but no answer text."]
        return "metadata_only", ["FAQ row has metadata but no question/answer content."]

    if content_type == "safety_summary":
        if title and body:
            return "full_content", ["Safety Summary has title and full guidance body."]
        if title and summary:
            return "content_brief", ["Safety Summary has concise title/summary guidance but no full content body."]
        if title:
            return "partial_content", ["Safety Summary has a title but incomplete guidance."]
        return "metadata_only", ["Safety Summary row has metadata but no clear guidance content."]

    return "malformed", ["Unknown content type."]


def locale_requirements(content_type, locale):
    suffix = "zh" if locale == "zh-CN" else locale
    if content_type == "resource":
        return [f"title_{suffix}", f"summary_{suffix}", f"body_{suffix}"]
    if content_type == "scenario":
        return [f"title_{suffix}", f"summary_{suffix}", "scenario_steps", "scenario_options"]
    if content_type == "faq":
        return [f"title_{suffix}", f"answer_{suffix}"]
    if content_type == "safety_summary":
        return [f"title_{suffix}", f"summary_{suffix}", f"body_{suffix}"]
    return []


def original_value(item, key):
    original = item.get("original", {})
    return original.get(key) or ""


def apply_locale_audit(item):
    content_type = item.get("normalized", {}).get("contentType")
    item["declaredLocaleCoverage"] = list(DECLARED_LOCALES)
    statuses = {}
    for locale in DECLARED_LOCALES:
        suffix = "zh" if locale == "zh-CN" else locale
        title = bool(original_value(item, f"title_{suffix}") or (locale == "en" and item.get("normalized", {}).get("title")))
        summary = bool(original_value(item, f"summary_{suffix}") or (locale == "en" and item.get("normalized", {}).get("summary")))
        body = bool(original_value(item, f"body_{suffix}") or original_value(item, f"answer_{suffix}") or (locale == "en" and item.get("normalized", {}).get("body")))

        if content_type == "scenario":
            has_structure = bool(item.get("scenarioStructure", {}).get("steps") and item.get("scenarioStructure", {}).get("options"))
            statuses[locale] = "complete" if title and summary and has_structure else ("partial" if title or summary else "missing")
        elif content_type in {"resource", "safety_summary"}:
            statuses[locale] = "complete" if title and summary and body else ("partial" if title or summary or body else "missing")
        elif content_type == "faq":
            statuses[locale] = "complete" if title and body else ("partial" if title or summary or body else "missing")
        else:
            statuses[locale] = "not_applicable"
    item["actualLocaleContent"] = statuses
    item["localeCompletenessRules"] = {
        "complete": "All meaningful learner-facing fields for this type and locale are present.",
        "partial": "Some localized learner-facing fields exist, but required content is incomplete.",
        "missing": "No meaningful localized learner-facing text was found.",
    }


def has_missing_content(item):
    completeness = item.get("contentCompleteness")
    if completeness in {"content_brief", "metadata_only", "partial_content"}:
        return True
    return bool(content_issue_codes(item) & MISSING_CONTENT_CODES)


def corrected_disposition(item):
    content_type = item.get("normalized", {}).get("contentType")
    mapping_method = item.get("taxonomyMapping", {}).get("mapping_method")
    duplicate_matches = item.get("duplicateMatches", [])
    exact_duplicate = any(match.get("recommendedDisposition") == "skip_exact_duplicate" for match in duplicate_matches)
    possible_duplicate = any(match.get("recommendedDisposition") == "possible_duplicate" or match.get("matchType") == "possible" for match in duplicate_matches)

    if content_type == "unsupported_content_type" or item.get("contentCompleteness") == "malformed":
        return "invalid_record"
    if content_type not in {"resource", "scenario", "faq", "safety_summary"}:
        return "unsupported_content_type"
    if mapping_method == "unresolved":
        return "needs_manual_mapping"
    if has_missing_content(item):
        return "missing_required_fields"
    if item.get("liveSupport") in {"planning_only", "unsupported"}:
        return "unsupported_live_type"
    if exact_duplicate:
        return "skip_exact_duplicate"
    if possible_duplicate:
        return "possible_duplicate"
    if item.get("contentCompleteness") == "full_content":
        return "create_candidate"
    return "missing_required_fields"


def import_readiness(item):
    if item.get("disposition") == "invalid_record":
        return "blocked_invalid_record"
    if item.get("taxonomyMapping", {}).get("mapping_method") == "unresolved":
        return "blocked_taxonomy_review"
    if has_missing_content(item) and item.get("liveSupport") == "live_supported":
        return "blocked_missing_content"
    if item.get("contentCompleteness") in {"content_brief", "metadata_only"}:
        return "planning_only"
    if item.get("liveSupport") in {"planning_only", "unsupported"}:
        return "blocked_unsupported_schema"
    if any(match.get("recommendedDisposition") in {"possible_duplicate", "skip_exact_duplicate"} for match in item.get("duplicateMatches", [])):
        return "blocked_duplicate_review"
    if item.get("contentCompleteness") == "full_content" and item.get("liveSupport") == "live_supported":
        return "ready_for_draft_import"
    return "planning_only"


def apply_audit_classification(item):
    completeness, reasons = classify_content_completeness(item)
    item["contentCompleteness"] = completeness
    item["contentCompletenessReasons"] = reasons
    apply_locale_audit(item)
    item["disposition"] = corrected_disposition(item)
    item["importReadiness"] = import_readiness(item)
    return item


def method_label(method):
    return {
        "content_id": "exact_content_id",
        "dry_run_id": "exact_dry_run_id",
        "slug": "exact_slug",
        "title": "exact_normalized_title",
        "fuzzy_title": "fuzzy_title",
    }.get(method or "", method or "unresolved")


def audit_relationship_resolution(relationship, resolution):
    source = resolution.get("source", {})
    target = resolution.get("target", {})
    source_method = method_label(source.get("method"))
    target_method = method_label(target.get("method"))
    uses_fuzzy = "fuzzy" in source_method or "fuzzy" in target_method
    review_required = "yes" if uses_fuzzy or source.get("status") != "resolved" or target.get("status") != "resolved" else "no"
    confidence = "high" if review_required == "no" and source_method == "exact_content_id" and target_method == "exact_content_id" else ("medium" if review_required == "no" else "manual_review_required")
    return {
        "relationship_id": relationship.get("originalRelationshipId"),
        "source_reference": relationship.get("sourceReference"),
        "target_reference": relationship.get("targetReference"),
        "relationship_type": relationship.get("relationshipType"),
        "source_resolution_method": source_method,
        "target_resolution_method": target_method,
        "source_resolved_id": source.get("dryRunId"),
        "target_resolved_id": target.get("dryRunId"),
        "confidence": confidence,
        "review_required": review_required,
        "note": resolution.get("ambiguityNotes") or "",
    }


def count_duplicate_records(items):
    duplicate_signals = sum(len(item.get("duplicateMatches", [])) for item in items)
    duplicate_records = sum(1 for item in items if item.get("duplicateMatches"))
    exact_duplicate_records = sum(1 for item in items if any(match.get("matchType") == "exact" for match in item.get("duplicateMatches", [])))
    possible_duplicate_records = sum(1 for item in items if any(match.get("recommendedDisposition") == "possible_duplicate" or match.get("matchType") == "possible" for match in item.get("duplicateMatches", [])))
    return {
        "duplicateSignals": duplicate_signals,
        "duplicateRecords": duplicate_records,
        "exactDuplicateRecords": exact_duplicate_records,
        "possibleDuplicateRecords": possible_duplicate_records,
    }


def find_header_row(rows, required):
    row_number, headers = preview.find_header(rows, required)
    return row_number, headers


def normalize_header_key(header):
    return preview.normalize_text(header).replace(" ", "_")


def extract_raw_row_structure(workbook_path):
    workbook_path = Path(workbook_path)
    rows = preview.read_sheet_rows(workbook_path, "02_Content_Inventory")
    header_row, headers = find_header_row(rows, ["content_code", "content_type", "topic_code"])
    mapped_headers = {
        "content_code", "content_type", "topic_code", "slug",
        "title_en", "title_ms", "title_zh", "summary_en", "summary_ms", "summary_zh",
        "difficulty_level", "estimated_minutes", "source_organisation", "source_url",
        "review_status", "rag_ready", "last_reviewed_at", "notes",
    }
    samples = {"resource": [], "scenario": [], "faq": [], "safety_summary": []}
    unused_columns = []
    if header_row:
        for index, header in enumerate(headers, start=1):
            if not header:
                continue
            key = normalize_header_key(header)
            if key in mapped_headers:
                continue
            sample_value = ""
            for values in rows[header_row:]:
                value = values[index - 1] if index - 1 < len(values) else None
                if value not in (None, ""):
                    sample_value = preview.value_to_json(value)
                    break
            if sample_value not in (None, ""):
                unused_columns.append({
                    "header": header,
                    "columnIndex": index,
                    "columnLetter": get_column_letter(index),
                    "sampleValue": sample_value,
                    "reasonNotMapped": "Governance/authoring metadata only; not learner-facing body, FAQ answer, scenario step, or option content.",
                })
        for offset, values in enumerate(rows[header_row:], start=header_row + 1):
            row = preview.row_to_dict(headers, values)
            content_code = str(preview.get_first(row, "content_code") or "").strip()
            if not content_code:
                continue
            content_type = preview.normalize_content_type(preview.get_first(row, "content_type"))
            if content_type in samples and len(samples[content_type]) < 2:
                samples[content_type].append({
                    "excelRow": offset,
                    "values": {key: preview.value_to_json(value) for key, value in row.items()},
                })
    relationship_rows = preview.read_sheet_rows(workbook_path, "03_Content_Relationships")
    rel_header_row, rel_headers = find_header_row(relationship_rows, ["relationship_id", "source_content_code", "target_content_code"])
    category_rows = preview.read_sheet_rows(workbook_path, "Cyberly_Content_Category_Plan")
    category_first_rows = []
    for row_number, values in enumerate(category_rows, start=1):
        visible_values = [preview.value_to_json(value) for value in values if value not in (None, "")]
        if visible_values:
            category_first_rows.append({"row": row_number, "values": visible_values})
        if len(category_first_rows) >= 10:
            break
    return {
        "contentInventory": {
            "headerRow": header_row,
            "headers": [{"header": header, "columnIndex": index, "columnLetter": get_column_letter(index)} for index, header in enumerate(headers, start=1) if header],
            "samples": samples,
            "unusedNonEmptyColumns": unused_columns,
        },
        "categoryPlan": {"sampleRows": category_first_rows},
        "contentRelationships": {
            "headerRow": rel_header_row,
            "headers": [{"header": header, "columnIndex": index, "columnLetter": get_column_letter(index)} for index, header in enumerate(rel_headers, start=1) if header],
        },
    }


def write_csv(path, rows, columns):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_text(path, lines):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def row_title(item):
    return item.get("normalized", {}).get("title") or item.get("dryRunId")


def duplicate_audit_rows(items):
    rows = []
    for item in items:
        for match in item.get("duplicateMatches", []):
            rows.append({
                "dryRunId": item["dryRunId"],
                "title": row_title(item),
                "contentType": item["normalized"]["contentType"],
                "matchedRecordId": match.get("matchedRecordId"),
                "matchedTitle": match.get("matchedTitle"),
                "signals": ", ".join(match.get("fieldsCompared", [])) if isinstance(match.get("fieldsCompared"), list) else match.get("fieldsCompared"),
                "matchType": match.get("matchType"),
                "recordDisposition": item.get("disposition"),
                "similarity": match.get("similarity", ""),
                "sourceUrlComparison": "matched" if "source_url" in match.get("fieldsCompared", []) else "not_matched",
                "slugComparison": "matched" if "slug" in match.get("fieldsCompared", []) else "not_matched",
                "titleComparison": "matched" if "title+category" in match.get("fieldsCompared", []) else "not_matched",
                "categoryComparison": "matched" if "title+category" in match.get("fieldsCompared", []) else "not_matched",
                "reason": match.get("reason"),
            })
    return rows


def build_audited_manifest(workbook_path, baseline_manifest_path, database_catalog_path, output_dir):
    output_dir = Path(output_dir)
    if not is_audit_output_path(output_dir):
        raise RuntimeError("Audit output must be written to a 14July2026-audit folder.")
    workbook_before = file_identity(workbook_path)
    baseline_manifest_identity = file_identity(baseline_manifest_path)
    baseline = json.loads(Path(baseline_manifest_path).read_text(encoding="utf-8"))
    catalog = json.loads(Path(database_catalog_path).read_text(encoding="utf-8"))
    assert_catalogue_read_only(catalog)
    raw_structure = extract_raw_row_structure(workbook_path)

    items = deepcopy(baseline.get("items", []))
    for item in items:
        apply_audit_classification(item)

    relationship_rows = []
    for relationship in baseline.get("relationships", []):
        relationship["resolutionAudit"] = audit_relationship_resolution(relationship, relationship.get("resolution", {}))
        relationship_rows.append(relationship["resolutionAudit"])

    duplicate_counts = count_duplicate_records(items)
    corrected_summary = {
        "candidateRecords": len(items),
        "byContentType": dict(sorted(Counter(item["normalized"]["contentType"] for item in items).items())),
        "byContentCompleteness": dict(sorted(Counter(item["contentCompleteness"] for item in items).items())),
        "byCorrectedDisposition": dict(sorted(Counter(item["disposition"] for item in items).items())),
        "byImportReadiness": dict(sorted(Counter(item["importReadiness"] for item in items).items())),
        "identityClearCandidates": sum(1 for item in items if not item.get("duplicateMatches")),
        "fullContentCandidates": sum(1 for item in items if item["contentCompleteness"] == "full_content"),
        "draftImportReadyCandidates": sum(1 for item in items if item["importReadiness"] == "ready_for_draft_import"),
        "duplicateAudit": duplicate_counts,
        "relationshipResolutionMethods": dict(sorted(Counter(
            row["source_resolution_method"] for row in relationship_rows
        ).items())),
        "relationshipTargetResolutionMethods": dict(sorted(Counter(
            row["target_resolution_method"] for row in relationship_rows
        ).items())),
        "relationshipReviewRequired": dict(sorted(Counter(row["review_required"] for row in relationship_rows).items())),
    }
    workbook_after = file_identity(workbook_path)
    if workbook_before != workbook_after:
        raise RuntimeError("Workbook identity changed during audit.")

    manifest = {
        "auditSchemaVersion": AUDIT_SCHEMA_VERSION,
        "mode": "dry-run-audit",
        "generatedAt": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "originalPhase2C1Manifest": baseline_manifest_identity,
        "workbook": workbook_before,
        "workbookAfterAudit": workbook_after,
        "database": catalog.get("database", {}),
        "correctedSummary": corrected_summary,
        "rawStructure": raw_structure,
        "items": items,
        "relationships": baseline.get("relationships", []),
        "sourceProvenance": baseline.get("items", []),
    }
    write_audit_artifacts(output_dir, manifest, raw_structure, relationship_rows)
    return manifest


def write_audit_artifacts(output_dir, manifest, raw_structure, relationship_rows):
    output_dir.mkdir(parents=True, exist_ok=True)
    preview.write_json_safe(output_dir / "content-import-manifest-audited.json", manifest)
    preview.write_json_safe(output_dir / "source-provenance-audited.json", {
        "workbook": manifest["workbook"],
        "originalPhase2C1Manifest": manifest["originalPhase2C1Manifest"],
        "items": [
            {
                "dryRunId": item["dryRunId"],
                "sourceSheet": item["provenance"]["sourceSheet"],
                "excelRow": item["provenance"]["excelRow"],
                "contentCompleteness": item["contentCompleteness"],
                "importReadiness": item["importReadiness"],
            }
            for item in manifest["items"]
        ],
    })
    preview.write_json_safe(output_dir / "raw-row-structure.json", raw_structure)

    rows = []
    locale_rows = []
    corrected_rows = []
    readiness_rows = []
    resource_rows = []
    faq_rows = []
    safety_rows = []
    for item in manifest["items"]:
        base = {
            "dry_run_id": item["dryRunId"],
            "content_type": item["normalized"]["contentType"],
            "title": row_title(item),
            "content_completeness": item["contentCompleteness"],
            "content_completeness_reasons": "; ".join(item["contentCompletenessReasons"]),
            "disposition": item["disposition"],
            "import_readiness": item["importReadiness"],
        }
        rows.append(base)
        corrected_rows.append({**base, "duplicate_signals": len(item.get("duplicateMatches", []))})
        readiness_rows.append({**base, "live_support": item.get("liveSupport")})
        for locale, status in item.get("actualLocaleContent", {}).items():
            locale_rows.append({
                "dry_run_id": item["dryRunId"],
                "content_type": item["normalized"]["contentType"],
                "title": row_title(item),
                "locale": locale,
                "declared": "yes" if locale in item.get("declaredLocaleCoverage", []) else "no",
                "actual_status": status,
            })
        if item["normalized"]["contentType"] == "resource":
            resource_rows.append({
                **base,
                "has_title": bool(item["normalized"].get("title")),
                "has_summary": bool(item["normalized"].get("summary")),
                "has_body": bool(item["normalized"].get("body")),
                "has_source_url": bool(item.get("sourceMetadata", {}).get("sourceUrl")),
                "body_availability": "no body field found in Content Inventory",
            })
        if item["normalized"]["contentType"] == "faq":
            faq_rows.append({
                **base,
                "has_question": bool(item["normalized"].get("title")),
                "has_answer": bool(item["normalized"].get("body")),
                "answer_availability": "no answer/response/guidance column found for FAQ rows",
            })
        if item["normalized"]["contentType"] == "safety_summary":
            safety_rows.append({
                **base,
                "has_title": bool(item["normalized"].get("title")),
                "has_summary": bool(item["normalized"].get("summary")),
                "has_body": bool(item["normalized"].get("body")),
                "guidance_availability": "concise summary present; no full body/guidance column found",
            })

    write_csv(output_dir / "content-completeness.csv", rows, [
        "dry_run_id", "content_type", "title", "content_completeness", "content_completeness_reasons", "disposition", "import_readiness",
    ])
    write_csv(output_dir / "locale-content-audit.csv", locale_rows, [
        "dry_run_id", "content_type", "title", "locale", "declared", "actual_status",
    ])
    write_csv(output_dir / "resource-structure-audit.csv", resource_rows, [
        "dry_run_id", "content_type", "title", "has_title", "has_summary", "has_body", "has_source_url", "body_availability", "content_completeness", "disposition", "import_readiness",
    ])
    write_csv(output_dir / "faq-structure-audit.csv", faq_rows, [
        "dry_run_id", "content_type", "title", "has_question", "has_answer", "answer_availability", "content_completeness", "disposition", "import_readiness",
    ])
    write_csv(output_dir / "safety-summary-structure-audit.csv", safety_rows, [
        "dry_run_id", "content_type", "title", "has_title", "has_summary", "has_body", "guidance_availability", "content_completeness", "disposition", "import_readiness",
    ])
    write_csv(output_dir / "relationship-resolution-audit.csv", relationship_rows, [
        "relationship_id", "source_reference", "target_reference", "relationship_type",
        "source_resolution_method", "target_resolution_method", "source_resolved_id",
        "target_resolved_id", "confidence", "review_required", "note",
    ])
    write_csv(output_dir / "corrected-dispositions.csv", corrected_rows, [
        "dry_run_id", "content_type", "title", "content_completeness", "disposition", "import_readiness", "duplicate_signals", "content_completeness_reasons",
    ])
    write_csv(output_dir / "import-readiness.csv", readiness_rows, [
        "dry_run_id", "content_type", "title", "live_support", "content_completeness", "disposition", "import_readiness", "content_completeness_reasons",
    ])

    write_duplicate_case_audit(output_dir / "duplicate-case-audit.md", manifest["items"])
    write_scenario_gap(output_dir / "scenario-structure-gap.md", manifest["items"])
    write_audit_summary(output_dir / "audit-summary.md", manifest)


def write_duplicate_case_audit(path, items):
    rows = duplicate_audit_rows(items)
    lines = [
        "# Duplicate Case Audit",
        "",
        "Duplicate signals are preserved independently from corrected disposition. A record can be incomplete and still keep its duplicate evidence.",
        "",
    ]
    if not rows:
        lines.append("No duplicate cases were detected.")
    for row in rows:
        lines.extend([
            f"## {row['dryRunId']}",
            f"- Workbook title: {row['title']}",
            f"- Content type: {row['contentType']}",
            f"- Matched live ID: {row['matchedRecordId']}",
            f"- Matched live title: {row['matchedTitle']}",
            f"- Match type: {row['matchType']}",
            f"- Matching signals: {row['signals']}",
            f"- Record-level disposition: {row['recordDisposition']}",
            f"- Source URL comparison: {row['sourceUrlComparison']}",
            f"- Slug comparison: {row['slugComparison']}",
            f"- Title/category comparison: {row['titleComparison']} / {row['categoryComparison']}",
            f"- Reason: {row['reason']}",
            "",
        ])
    write_text(path, lines)


def write_scenario_gap(path, items):
    scenarios = [item for item in items if item["normalized"]["contentType"] == "scenario"]
    lines = [
        "# Scenario Structure Gap",
        "",
        "| Workbook field | Current live schema field | Available | Mapping | Gap | Import consequence |",
        "| --- | --- | --- | --- | --- | --- |",
        "| title_en/ms/zh | scenario_definitions.title + translations | Yes | direct title text | no step context | title only is insufficient |",
        "| summary_en/ms/zh | scenario_definitions.summary + translations | Yes | direct summary text | no scenario step text | description only is insufficient |",
        "| difficulty_level | scenario_definitions.difficulty | Yes | needs level mapping review | values use L1-style levels | manual mapping required |",
        "| estimated_minutes | scenario_definitions.estimated_minutes | Yes | direct duration | none | usable metadata |",
        "| scenario steps | scenario_steps.situation_text/prompt_text | No | none | absent from workbook row | cannot create live Scenario |",
        "| options | scenario_steps.options_json | No | none | absent from workbook row | cannot create live Scenario |",
        "| feedback/scoring | scenario options feedback/scoring | No | none | absent from workbook row | cannot create live Scenario |",
        "",
        f"Scenario rows audited: {len(scenarios)}. All are content briefs until step and option structures are authored.",
    ]
    write_text(path, lines)


def write_audit_summary(path, manifest):
    summary = manifest["correctedSummary"]
    lines = [
        "# Phase 2C.2 Dry-Run Audit Summary",
        "",
        "This audit corrects Phase 2C.1 terminology. The workbook rows are planning records and authoring briefs, not import-ready production content.",
        "",
        "## Key Answers",
        f"- Resources ready for draft import: {summary['byImportReadiness'].get('ready_for_draft_import', 0)}",
        f"- Scenarios ready for draft import: 0",
        "- FAQ answers actually present: no, not in the detected FAQ answer/body/guidance fields.",
        "- Safety Summaries: concise briefs, not full content bodies.",
        "- Locale coverage: declared workbook columns exist, but actual completeness is partial because full bodies/answers/steps/options are missing.",
        "- Phase 2C.1 called 18 records create candidates because it treated non-duplicate live-supported records as candidates before considering full content completeness.",
        f"- Actually ready now: {summary['draftImportReadyCandidates']}",
        "- Authoring backlog: Resource bodies, Scenario steps/options, FAQ answers, and full Safety Summary guidance.",
        f"- Duplicate records needing review or skip handling: {summary['duplicateAudit']['duplicateRecords']}",
        f"- Relationship trust: {summary['relationshipReviewRequired'].get('no', 0)} exact-ID resolved, {summary['relationshipReviewRequired'].get('yes', 0)} require review.",
        "",
        "## Corrected Counts",
        f"- By completeness: `{json.dumps(summary['byContentCompleteness'], sort_keys=True)}`",
        f"- By corrected disposition: `{json.dumps(summary['byCorrectedDisposition'], sort_keys=True)}`",
        f"- By import readiness: `{json.dumps(summary['byImportReadiness'], sort_keys=True)}`",
    ]
    write_text(path, lines)


def main(argv=None):
    parser = argparse.ArgumentParser(description="Audit Cyberly workbook dry-run manifest without modifying source artifacts.")
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--baseline-manifest", required=True)
    parser.add_argument("--database-catalog", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args(argv)
    manifest = build_audited_manifest(args.workbook, args.baseline_manifest, args.database_catalog, args.output)
    print(json.dumps({
        "output": args.output,
        "originalManifestHash": manifest["originalPhase2C1Manifest"]["sha256"],
        "workbookHash": manifest["workbook"]["sha256"],
        "correctedDisposition": manifest["correctedSummary"]["byCorrectedDisposition"],
        "importReadiness": manifest["correctedSummary"]["byImportReadiness"],
        "databaseUnchanged": manifest["database"].get("unchanged"),
    }, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
