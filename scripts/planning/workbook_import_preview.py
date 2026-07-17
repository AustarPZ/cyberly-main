import argparse
import csv
import hashlib
import json
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from copy import deepcopy
from datetime import date, datetime, timezone
from difflib import SequenceMatcher
from pathlib import Path
from urllib.parse import urlparse, urlunparse

try:
    import openpyxl
except ImportError:  # pragma: no cover - exercised by CLI environment
    openpyxl = None


MAIN_CATEGORIES = [
    {"name": "Beginner / Digital Foundations", "liveCode": "Beginner"},
    {"name": "Scams & Social Engineering", "liveCode": "Scams"},
    {"name": "Passwords & Account Security", "liveCode": "Passwords"},
    {"name": "Privacy & Personal Data Protection", "liveCode": "Privacy"},
    {"name": "Online Safety & Digital Wellbeing", "liveCode": "Safety"},
    {"name": "Misinformation & Media Literacy", "liveCode": "Misinformation"},
    {"name": "AI & Technology Safety", "liveCode": "AI & Technology"},
]

ALIAS_TO_CATEGORY = {
    "topic scams": "Scams & Social Engineering",
    "scams": "Scams & Social Engineering",
    "scams social engineering": "Scams & Social Engineering",
    "phishing and scams": "Scams & Social Engineering",
    "misinformation": "Misinformation & Media Literacy",
    "topic misinformation": "Misinformation & Media Literacy",
    "misinformation media literacy": "Misinformation & Media Literacy",
    "ai and technology safety": "AI & Technology Safety",
    "ai technology safety": "AI & Technology Safety",
    "topic ai technology": "AI & Technology Safety",
    "topic ai and technology": "AI & Technology Safety",
    "privacy": "Privacy & Personal Data Protection",
    "topic privacy": "Privacy & Personal Data Protection",
    "privacy personal data protection": "Privacy & Personal Data Protection",
    "online safety": "Online Safety & Digital Wellbeing",
    "topic safety": "Online Safety & Digital Wellbeing",
    "online safety digital wellbeing": "Online Safety & Digital Wellbeing",
    "password safety": "Passwords & Account Security",
    "topic password": "Passwords & Account Security",
    "passwords account security": "Passwords & Account Security",
    "beginner digital foundations": "Beginner / Digital Foundations",
    "topic beginner": "Beginner / Digital Foundations",
    "beginner digital foundation": "Beginner / Digital Foundations",
}

SUBCATEGORY_TO_CATEGORY = {
    "digital citizenship": "Beginner / Digital Foundations",
    "social engineering basics": "Beginner / Digital Foundations",
    "safe digital habits": "Beginner / Digital Foundations",
    "think before clicking": "Beginner / Digital Foundations",
    "phishing": "Scams & Social Engineering",
    "qr code scams": "Scams & Social Engineering",
    "sim swap otp fraud": "Scams & Social Engineering",
    "online shopping scams": "Scams & Social Engineering",
    "strong passwords": "Passwords & Account Security",
    "multi factor authentication": "Passwords & Account Security",
    "password managers": "Passwords & Account Security",
    "suspicious login alerts": "Passwords & Account Security",
    "personal data": "Privacy & Personal Data Protection",
    "app permissions": "Privacy & Personal Data Protection",
    "social media privacy": "Privacy & Personal Data Protection",
    "doxxing": "Privacy & Personal Data Protection",
    "online gaming safety": "Online Safety & Digital Wellbeing",
    "cyberbullying": "Online Safety & Digital Wellbeing",
    "digital wellbeing": "Online Safety & Digital Wellbeing",
    "fake news": "Misinformation & Media Literacy",
    "deepfakes": "Misinformation & Media Literacy",
    "media literacy": "Misinformation & Media Literacy",
    "ai generated scam messages": "AI & Technology Safety",
    "ai chatbot privacy": "AI & Technology Safety",
}

CONTENT_TYPE_ALIASES = {
    "resource": "resource",
    "resources": "resource",
    "scenario": "scenario",
    "scenarios": "scenario",
    "faq": "faq",
    "faqs": "faq",
    "safety summary": "safety_summary",
    "safety summaries": "safety_summary",
    "safety_summary": "safety_summary",
}

SUPPORTED_RELATIONSHIP_TYPES = {
    "prerequisite",
    "next_step",
    "practice_after",
    "remedial",
    "related_content",
    "related_topic",
    "prepare_before",
}

SECRET_KEY_PATTERN = re.compile(r"(password|secret|api[_-]?key|token|session)", re.I)


def normalize_text(value):
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = text.replace("&", " and ")
    text = re.sub(r"[^A-Za-z0-9\u4e00-\u9fff]+", " ", text)
    return re.sub(r"\s+", " ", text).strip().lower()


def category_code_for_name(name):
    for category in MAIN_CATEGORIES:
        if category["name"] == name:
            return category["liveCode"]
    return ""


def map_topic(topic):
    original = "" if topic is None else str(topic).strip()
    normalized = normalize_text(original)
    exact_names = {normalize_text(row["name"]): row["name"] for row in MAIN_CATEGORIES}
    exact_codes = {normalize_text(row["liveCode"]): row["name"] for row in MAIN_CATEGORIES}

    mapped = None
    method = "unresolved"
    proposed_subcategory = ""
    confidence = "manual_review_required"
    note = "Manual taxonomy mapping is required."

    if normalized in exact_names:
        mapped = exact_names[normalized]
        method = "exact_category"
        confidence = "high"
        note = "Matched current Cyberly category label exactly."
    elif normalized in exact_codes:
        mapped = exact_codes[normalized]
        method = "alias_exact"
        confidence = "high"
        note = "Matched current live category code."
    elif normalized in ALIAS_TO_CATEGORY:
        mapped = ALIAS_TO_CATEGORY[normalized]
        method = "alias_exact"
        confidence = "high"
        note = "Matched deterministic legacy topic alias."
    elif normalized in SUBCATEGORY_TO_CATEGORY:
        mapped = SUBCATEGORY_TO_CATEGORY[normalized]
        method = "subcategory_exact"
        proposed_subcategory = original
        confidence = "medium"
        note = "Matched a known planning subcategory; live subcategory support is not available."
    elif "phish" in normalized or "scam" in normalized:
        mapped = "Scams & Social Engineering"
        method = "deterministic_title_rule"
        confidence = "medium"
        proposed_subcategory = original if original else "Phishing"
        note = "Matched deterministic phishing/scam keyword rule."

    return {
        "original_topic": original,
        "normalized_topic": normalized,
        "mapped_category": mapped or "",
        "live_category_code": category_code_for_name(mapped) if mapped else "",
        "proposed_subcategory": proposed_subcategory,
        "mapping_method": method,
        "confidence": confidence,
        "manual_review_note": note,
    }


def normalize_content_type(value):
    return CONTENT_TYPE_ALIASES.get(normalize_text(value).replace(" ", "_"), CONTENT_TYPE_ALIASES.get(normalize_text(value), "unsupported_content_type"))


def slugify(title):
    ascii_text = unicodedata.normalize("NFKD", "" if title is None else str(title)).encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^A-Za-z0-9]+", "-", ascii_text.lower()).strip("-")
    return re.sub(r"-+", "-", slug)


def canonical_url(value):
    if not value:
        return ""
    text = str(value).strip()
    parsed = urlparse(text)
    if not parsed.scheme or not parsed.netloc:
        return ""
    path = re.sub(r"/+$", "", parsed.path)
    return urlunparse((parsed.scheme.lower(), parsed.netloc.lower(), path, "", "", ""))


def row_to_dict(headers, values):
    row = {}
    for index, header in enumerate(headers):
        if not header:
            continue
        value = values[index] if index < len(values) else None
        row[str(header).strip()] = value
    return row


def value_to_json(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if value is None:
        return ""
    return value


def get_first(row, *names):
    lowered = {normalize_text(key).replace(" ", "_"): value for key, value in row.items()}
    for name in names:
        key = normalize_text(name).replace(" ", "_")
        if key in lowered:
            return lowered[key]
    return ""


def load_catalog(path):
    with open(path, "r", encoding="utf-8") as handle:
        return json.load(handle)


def sha256_file(path):
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest().upper()


def workbook_identity(path):
    stat = Path(path).stat()
    return {
        "path": str(Path(path).as_posix()),
        "fileName": Path(path).name,
        "sha256": sha256_file(path),
        "sizeBytes": stat.st_size,
        "lastModified": datetime.fromtimestamp(stat.st_mtime).isoformat(),
        "authoritativeVersion": "14July2026" if "14July2026" in Path(path).name else "custom",
    }


def inspect_workbook(path):
    if openpyxl is None:
        raise RuntimeError("openpyxl is unavailable. Use an isolated local environment with openpyxl installed.")
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = []
        for worksheet in workbook.worksheets:
            rows_seen = 0
            max_cols = 0
            non_empty_rows = 0
            header_rows = []
            blank_inside = 0
            seen_data = False
            duplicate_headers = []
            for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                rows_seen = row_number
                max_cols = max(max_cols, len(row))
                values = [cell for cell in row if cell not in (None, "")]
                if values:
                    non_empty_rows += 1
                    seen_data = True
                    normalized_headers = [normalize_text(cell) for cell in row if cell not in (None, "")]
                    if len(normalized_headers) >= 3:
                        dupes = [item for item, count in Counter(normalized_headers).items() if count > 1]
                        if dupes:
                            duplicate_headers.extend(dupes)
                    if any(str(cell).strip().lower() in {"content_code", "relationship_id"} for cell in row if cell is not None):
                        header_rows.append({"row": row_number, "headers": [str(cell) for cell in row if cell not in (None, "")]})
                elif seen_data:
                    blank_inside += 1
            sheets.append({
                "name": worksheet.title,
                "visibility": worksheet.sheet_state,
                "dimensions": {"rows": rows_seen, "columns": max_cols, "nonEmptyRows": non_empty_rows},
                "headerRows": header_rows,
                "blankRowsInsideDataRanges": blank_inside,
                "duplicateHeaders": sorted(set(duplicate_headers)),
                "mergedCells": "not available in read_only mode",
                "hiddenRowsColumns": "not inspected in read_only mode",
                "formulaCellsWithoutCachedValues": "not inspected in data_only read_only mode",
            })
        return sheets
    finally:
        workbook.close()


def find_header(rows, required):
    required_norm = {normalize_text(item).replace(" ", "_") for item in required}
    for index, row in enumerate(rows, start=1):
        headers = [str(cell).strip() if cell is not None else "" for cell in row]
        normalized = {normalize_text(header).replace(" ", "_") for header in headers}
        if required_norm.issubset(normalized):
            return index, headers
    return None, []


def read_sheet_rows(workbook_path, sheet_name):
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            return []
        return list(workbook[sheet_name].iter_rows(values_only=True))
    finally:
        workbook.close()


def build_candidate(row, sheet, excel_row):
    content_id = get_first(row, "content_code", "id")
    raw_type = get_first(row, "content_type")
    content_type = normalize_content_type(raw_type)
    title = get_first(row, "title_en", "title", "question_en")
    summary = get_first(row, "summary_en", "description_en", "answer_en")
    body = get_first(row, "body_en", "content_en", "guidance_en", "answer_en")
    topic = get_first(row, "topic_code", "topic", "category")
    mapping = map_topic(topic)
    explicit_slug = get_first(row, "slug")
    proposed_slug = str(explicit_slug).strip() if explicit_slug else slugify(title)
    source_url = get_first(row, "source_url", "url")
    source_org = get_first(row, "source_organisation", "source_organization", "source_label")
    dry_run_id = f"workbook:{sheet}:{excel_row}"
    locale_coverage = {
        "en": bool(title or summary or body),
        "ms": bool(get_first(row, "title_ms", "summary_ms", "body_ms", "answer_ms")),
        "zh-CN": bool(get_first(row, "title_zh", "summary_zh", "body_zh", "answer_zh")),
    }
    candidate = {
        "dryRunId": dry_run_id,
        "provenance": {"sourceSheet": sheet, "excelRow": excel_row, "workbookContentId": value_to_json(content_id)},
        "original": {key: value_to_json(value) for key, value in row.items()},
        "normalized": {
            "contentType": content_type,
            "title": "" if title is None else str(title).strip(),
            "summary": "" if summary is None else str(summary).strip(),
            "body": "" if body is None else str(body).strip(),
            "difficultyLevel": value_to_json(get_first(row, "difficulty_level", "difficulty")),
            "estimatedMinutes": value_to_json(get_first(row, "estimated_minutes", "duration_minutes")),
        },
        "taxonomyMapping": mapping,
        "proposedSlug": proposed_slug,
        "localeCoverage": locale_coverage,
        "sourceMetadata": {
            "sourceLabel": "" if source_org is None else str(source_org).strip(),
            "sourceOrganisation": "" if source_org is None else str(source_org).strip(),
            "sourceUrl": "" if source_url is None else str(source_url).strip(),
            "canonicalSourceUrl": canonical_url(source_url),
            "sourceType": value_to_json(get_first(row, "source_type")),
            "sourceCountry": value_to_json(get_first(row, "source_country")),
            "sourceAuthorityLevel": value_to_json(get_first(row, "source_authority_level")),
        },
        "safetyMetadata": {
            "ageSuitability": value_to_json(get_first(row, "age_suitability", "age_appropriateness")),
            "sensitiveTopicFlag": value_to_json(get_first(row, "sensitive_topic_flag")),
            "malaysiaGuidanceFlag": value_to_json(get_first(row, "malaysia_guidance_flag")),
        },
        "workbookGovernance": {
            "workbookReviewStatus": value_to_json(get_first(row, "review_status")),
            "workbookRagReady": value_to_json(get_first(row, "rag_ready")),
            "workbookSourceChecked": value_to_json(get_first(row, "last_reviewed_at")),
            "workbookApprovalNotes": value_to_json(get_first(row, "notes")),
        },
        "proposedSafeDefaults": {
            "publicationStatus": "draft",
            "reviewStatus": "draft",
            "ragReady": False,
            "effectiveRagEligible": False,
        },
        "scenarioStructure": {
            "steps": int(float(get_first(row, "number_of_decisions", "steps") or 0)) if str(get_first(row, "number_of_decisions", "steps") or "").replace(".", "", 1).isdigit() else 0,
            "options": int(float(get_first(row, "options") or 0)) if str(get_first(row, "options") or "").replace(".", "", 1).isdigit() else 0,
        },
        "validation": {"status": "pending", "issues": []},
        "duplicateMatches": [],
        "liveSupport": "unsupported",
        "disposition": "invalid_record",
        "manualReviewReasons": [],
    }
    if not candidate["proposedSlug"] and candidate["normalized"]["title"]:
        candidate["manualReviewReasons"].append("manual_slug_required")
    return candidate


def extract_content_candidates(workbook_path):
    rows = read_sheet_rows(workbook_path, "02_Content_Inventory")
    header_row, headers = find_header(rows, ["content_code", "content_type", "topic_code"])
    candidates = []
    if not header_row:
        return candidates, ["Content Inventory header row was not found."]
    for offset, values in enumerate(rows[header_row:], start=header_row + 1):
        if not any(cell not in (None, "") for cell in values):
            continue
        row = row_to_dict(headers, values)
        raw_type = get_first(row, "content_type")
        content_code = str(get_first(row, "content_code") or "").strip()
        if not raw_type or not re.match(r"^(RES|SCN|FAQ|SAFE)-", content_code, re.I):
            continue
        candidates.append(build_candidate(row, "02_Content_Inventory", offset))
    return candidates, []


def extract_relationships(workbook_path):
    rows = read_sheet_rows(workbook_path, "03_Content_Relationships")
    header_row, headers = find_header(rows, ["relationship_id", "source_content_code", "target_content_code"])
    relationships = []
    if not header_row:
        return relationships, ["Content Relationships header row was not found."]
    for offset, values in enumerate(rows[header_row:], start=header_row + 1):
        if not any(cell not in (None, "") for cell in values):
            continue
        row = row_to_dict(headers, values)
        relationship_id = str(get_first(row, "relationship_id") or "").strip()
        if not re.match(r"^REL-", relationship_id, re.I):
            continue
        rel_type = normalize_text(get_first(row, "relationship_type")).replace(" ", "_")
        relationships.append({
            "dryRunRelationshipId": f"workbook:03_Content_Relationships:{offset}",
            "sourceSheet": "03_Content_Relationships",
            "excelRow": offset,
            "originalRelationshipId": value_to_json(relationship_id),
            "sourceReference": value_to_json(get_first(row, "source_content_code", "source")),
            "targetReference": value_to_json(get_first(row, "target_content_code", "target")),
            "relationshipType": rel_type,
            "topic": value_to_json(get_first(row, "topic_code")),
            "priority": value_to_json(get_first(row, "priority")),
            "sequenceOrder": value_to_json(get_first(row, "sequence_order")),
            "resolution": {},
        })
    return relationships, []


def validate_candidate(candidate):
    content_type = candidate.get("normalized", {}).get("contentType")
    normalized = candidate.get("normalized", {})
    source = candidate.get("sourceMetadata", {})
    issues = []

    def add(severity, field, code, message, action):
        issues.append({"severity": severity, "field": field, "issue_code": code, "message": message, "recommended_action": action})

    if content_type == "unsupported_content_type":
        add("error", "content_type", "unsupported_content_type", "Workbook content type is not supported for this dry-run scope.", "Review manually.")
        return issues

    if not normalized.get("title"):
        add("error", "title", "missing_title", "Title is required.", "Add a reviewed title before import.")
    if not candidate.get("taxonomyMapping", {}).get("live_category_code"):
        add("error", "topic", "unmapped_topic", "Topic cannot be mapped to a live Cyberly category.", "Map the topic manually.")
    source_url = source.get("sourceUrl")
    if source_url and not canonical_url(source_url):
        add("warning", "source_url", "malformed_source_url", "Source URL is not a valid absolute URL.", "Verify or replace the source URL.")

    if content_type == "resource":
        if not normalized.get("summary"):
            add("error", "summary", "missing_summary", "Resource summary is required.", "Add a reviewed summary.")
        if not normalized.get("body"):
            add("warning", "body", "missing_body", "Workbook does not include Resource body/content text.", "Provide body content before import.")
        if not source_url:
            add("warning", "source_url", "missing_source_url", "Resource source URL is missing.", "Add a source URL before review.")
    elif content_type == "scenario":
        if not normalized.get("summary"):
            add("error", "description", "missing_description", "Scenario description is required.", "Add a scenario description.")
        structure = candidate.get("scenarioStructure", {})
        if not structure.get("steps"):
            add("warning", "scenario_steps", "missing_scenario_steps", "Scenario step structure is not present in the workbook row.", "Model scenario steps before import.")
        if not structure.get("options"):
            add("warning", "scenario_options", "missing_scenario_options", "Scenario options/decisions are not present in the workbook row.", "Model decision options before import.")
    elif content_type == "faq":
        if not normalized.get("body"):
            add("error", "answer", "missing_faq_answer", "FAQ answer is missing.", "Add a reviewed FAQ answer.")
    elif content_type == "safety_summary":
        if not normalized.get("summary") and not normalized.get("body"):
            add("error", "guidance", "missing_safety_guidance", "Safety Summary guidance is missing.", "Add concise reviewed guidance.")

    return issues


def catalog_items_for_type(catalog, content_type):
    if content_type == "resource":
        return catalog.get("resources", [])
    if content_type == "scenario":
        return catalog.get("scenarios", [])
    return []


def live_item_id(item, content_type):
    prefix = "resource" if content_type == "resource" else "scenario"
    return f"live:{prefix}:{item.get('id')}"


def detect_live_duplicates(candidate, catalog):
    content_type = candidate.get("normalized", {}).get("contentType")
    if content_type not in {"resource", "scenario"}:
        return []
    matches = []
    title_norm = normalize_text(candidate.get("normalized", {}).get("title"))
    source_url = candidate.get("sourceMetadata", {}).get("canonicalSourceUrl")
    slug = normalize_text(candidate.get("proposedSlug")).replace(" ", "-")
    category = candidate.get("taxonomyMapping", {}).get("live_category_code")
    seen_live = set()
    for item in catalog_items_for_type(catalog, content_type):
        item_record_id = live_item_id(item, content_type)
        if item_record_id in seen_live:
            continue
        item_title = item.get("title") or ""
        item_title_norm = normalize_text(item_title)
        item_slug = normalize_text(item.get("slug")).replace(" ", "-")
        item_category = item.get("category_code") or item.get("topic_code") or ""
        item_source = canonical_url(item.get("source_url"))
        fields = []
        if slug and item_slug and slug == item_slug:
            fields.append("slug")
        if title_norm and item_title_norm and title_norm == item_title_norm and normalize_text(category) == normalize_text(item_category):
            fields.append("title+category")
        if source_url and item_source and source_url == item_source:
            fields.append("source_url")
        if fields:
            matches.append({
                "matchType": "exact",
                "workbookRecordId": candidate["dryRunId"],
                "matchedRecordId": item_record_id,
                "matchedTitle": item_title,
                "reason": f"Exact match on {', '.join(fields)}.",
                "fieldsCompared": fields,
                "confidence": "high",
                "recommendedDisposition": "skip_exact_duplicate" if "slug" in fields or "title+category" in fields else "possible_duplicate",
            })
            seen_live.add(item_record_id)
            continue
        if title_norm and item_title_norm and normalize_text(category) == normalize_text(item_category):
            score = SequenceMatcher(None, title_norm, item_title_norm).ratio()
            if score >= 0.86:
                matches.append({
                    "matchType": "possible",
                    "workbookRecordId": candidate["dryRunId"],
                    "matchedRecordId": item_record_id,
                    "matchedTitle": item_title,
                    "reason": f"Title similarity {score:.2f} using difflib SequenceMatcher threshold 0.86.",
                    "fieldsCompared": ["title", "category"],
                    "confidence": "medium",
                    "recommendedDisposition": "possible_duplicate",
                })
                seen_live.add(item_record_id)
    return matches


def detect_workbook_duplicates(candidates):
    matches = defaultdict(list)
    seen_slugs = {}
    seen_titles = {}
    for candidate in candidates:
        content_type = candidate["normalized"]["contentType"]
        category = candidate["taxonomyMapping"]["live_category_code"]
        slug = candidate.get("proposedSlug")
        title_key = (content_type, category, normalize_text(candidate["normalized"]["title"]))
        if slug:
            slug_key = normalize_text(slug)
            if slug_key in seen_slugs:
                other = seen_slugs[slug_key]
                match = {
                    "matchType": "exact",
                    "workbookRecordId": candidate["dryRunId"],
                    "matchedRecordId": other["dryRunId"],
                    "matchedTitle": other["normalized"]["title"],
                    "reason": "Duplicate workbook slug.",
                    "fieldsCompared": ["slug"],
                    "confidence": "high",
                    "recommendedDisposition": "possible_duplicate",
                }
                matches[candidate["dryRunId"]].append(match)
            else:
                seen_slugs[slug_key] = candidate
        if title_key[2]:
            if title_key in seen_titles:
                other = seen_titles[title_key]
                matches[candidate["dryRunId"]].append({
                    "matchType": "exact",
                    "workbookRecordId": candidate["dryRunId"],
                    "matchedRecordId": other["dryRunId"],
                    "matchedTitle": other["normalized"]["title"],
                    "reason": "Duplicate workbook title within same type and category.",
                    "fieldsCompared": ["title", "content_type", "category"],
                    "confidence": "high",
                    "recommendedDisposition": "possible_duplicate",
                })
            else:
                seen_titles[title_key] = candidate
    return matches


def find_slug_collisions(candidates, live_items):
    collisions = []
    live_slugs = {normalize_text(item.get("slug")): item for item in live_items if item.get("slug")}
    seen = {}
    for candidate in candidates:
        slug = candidate.get("proposedSlug")
        if not slug:
            continue
        key = normalize_text(slug)
        if key in live_slugs:
            collisions.append({"dryRunId": candidate["dryRunId"], "slug": slug, "matched": f"live:{live_slugs[key].get('id')}"})
        elif key in seen:
            collisions.append({"dryRunId": candidate["dryRunId"], "slug": slug, "matched": seen[key]})
        else:
            seen[key] = candidate["dryRunId"]
    return collisions


def live_support_for(content_type, catalog):
    support = catalog.get("support", {})
    if content_type == "resource":
        return support.get("resource", "live_supported")
    if content_type == "scenario":
        return support.get("scenario", "live_supported")
    if content_type == "faq":
        return support.get("faq", "planning_only")
    if content_type == "safety_summary":
        return support.get("safetySummary", "planning_only")
    return "unsupported"


def classify_disposition(candidate):
    content_type = candidate["normalized"]["contentType"]
    issues = candidate["validation"]["issues"]
    blocking_errors = [issue for issue in issues if issue["severity"] == "error"]
    exact = [match for match in candidate["duplicateMatches"] if match["matchType"] == "exact" and match["recommendedDisposition"] == "skip_exact_duplicate"]
    possible = [match for match in candidate["duplicateMatches"] if match["matchType"] == "possible" or match["recommendedDisposition"] == "possible_duplicate"]
    if content_type == "unsupported_content_type":
        return "invalid_record"
    if candidate["taxonomyMapping"]["mapping_method"] == "unresolved":
        return "needs_manual_mapping"
    if candidate["liveSupport"] in {"planning_only", "unsupported"}:
        return "unsupported_live_type"
    if blocking_errors:
        return "missing_required_fields"
    if exact:
        return "skip_exact_duplicate"
    if possible:
        return "possible_duplicate"
    return "create_candidate"


def build_candidate_index(candidates):
    index = {"byId": defaultdict(list), "byTitle": defaultdict(list), "bySlug": defaultdict(list), "byDryRunId": {}}
    for candidate in candidates:
        index["byDryRunId"][candidate["dryRunId"]] = candidate
        content_id = normalize_text(
            candidate.get("original", {}).get("content_code")
            or candidate.get("original", {}).get("contentId")
            or candidate.get("provenance", {}).get("workbookContentId")
        )
        title = normalize_text(candidate.get("normalized", {}).get("title"))
        slug = normalize_text(candidate.get("proposedSlug"))
        if content_id:
            index["byId"][content_id].append(candidate)
        if title:
            index["byTitle"][title].append(candidate)
        if slug:
            index["bySlug"][slug].append(candidate)
    return index


def resolve_relationship_endpoint(reference, index):
    key = normalize_text(reference)
    if not key:
        return {"status": "unresolved", "method": "", "dryRunId": "", "note": "Blank reference."}
    for method, bucket in [("content_id", "byId"), ("title", "byTitle"), ("slug", "bySlug")]:
        matches = index[bucket].get(key, [])
        if len(matches) == 1:
            return {"status": "resolved", "method": method, "dryRunId": matches[0]["dryRunId"], "note": ""}
        if len(matches) > 1:
            return {"status": "ambiguous", "method": method, "dryRunId": "", "note": f"{len(matches)} candidates matched."}
    return {"status": "unresolved", "method": "", "dryRunId": "", "note": "No exact ID, title, or slug match."}


def resolve_relationships(relationships, candidates):
    index = build_candidate_index(candidates)
    for relationship in relationships:
        source = resolve_relationship_endpoint(relationship["sourceReference"], index)
        target = resolve_relationship_endpoint(relationship["targetReference"], index)
        rel_type = relationship["relationshipType"]
        if rel_type not in SUPPORTED_RELATIONSHIP_TYPES:
            disposition = "unsupported_relationship_type"
        elif source["status"] == "unresolved":
            disposition = "unresolved_source"
        elif source["status"] == "ambiguous":
            disposition = "ambiguous_source"
        elif target["status"] == "unresolved":
            disposition = "unresolved_target"
        elif target["status"] == "ambiguous":
            disposition = "ambiguous_target"
        else:
            disposition = "resolved"
        relationship["resolution"] = {
            "source": source,
            "target": target,
            "disposition": disposition,
            "ambiguityNotes": "; ".join(note for note in [source.get("note"), target.get("note")] if note),
        }
    return relationships


def summarize(candidates, relationships, issues):
    by_type = Counter(candidate["normalized"]["contentType"] for candidate in candidates)
    by_disposition = Counter(candidate["disposition"] for candidate in candidates)
    exact = sum(1 for candidate in candidates for match in candidate["duplicateMatches"] if match["matchType"] == "exact")
    possible = sum(1 for candidate in candidates for match in candidate["duplicateMatches"] if match["recommendedDisposition"] == "possible_duplicate")
    validation_errors = sum(1 for candidate in candidates for issue in candidate["validation"]["issues"] if issue["severity"] == "error")
    return {
        "totalRows": len(candidates),
        "contentCandidates": len(candidates),
        "byContentType": dict(sorted(by_type.items())),
        "byDisposition": dict(sorted(by_disposition.items())),
        "unresolvedMappings": sum(1 for candidate in candidates if candidate["taxonomyMapping"]["mapping_method"] == "unresolved"),
        "exactDuplicates": exact,
        "possibleDuplicates": possible,
        "validationErrors": validation_errors,
        "relationshipDispositions": dict(sorted(Counter(rel["resolution"]["disposition"] for rel in relationships).items())),
        "structuralIssues": len(issues),
    }


def sanitize_for_json(value):
    if isinstance(value, dict):
        return {key: sanitize_for_json(inner) for key, inner in value.items() if not SECRET_KEY_PATTERN.search(str(key))}
    if isinstance(value, list):
        return [sanitize_for_json(item) for item in value]
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def write_json_safe(path, payload):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(sanitize_for_json(payload), ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def write_csv(path, rows, columns):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column, "") for column in columns})


def extract_category_plan(workbook_path):
    rows = read_sheet_rows(workbook_path, "Cyberly_Content_Category_Plan")
    non_empty = []
    for row_number, row in enumerate(rows, start=1):
        values = [value_to_json(cell) for cell in row if cell not in (None, "")]
        if values:
            non_empty.append({"row": row_number, "values": values})
    return non_empty


def build_taxonomy_rows(candidates):
    rows = {}
    for candidate in candidates:
        mapping = candidate["taxonomyMapping"]
        rows[mapping["original_topic"] or candidate["dryRunId"]] = mapping
    return [rows[key] for key in sorted(rows)]


def process(workbook_path, catalog_path, output_dir):
    workbook_path = Path(workbook_path)
    catalog_path = Path(catalog_path)
    output_dir = Path(output_dir)
    before_identity = workbook_identity(workbook_path)
    catalog = load_catalog(catalog_path)
    sheet_inventory = inspect_workbook(workbook_path)
    candidates, candidate_warnings = extract_content_candidates(workbook_path)
    relationships, relationship_warnings = extract_relationships(workbook_path)
    category_plan = extract_category_plan(workbook_path)

    workbook_dupes = detect_workbook_duplicates(candidates)
    live_items = list(catalog.get("resources", [])) + list(catalog.get("scenarios", []))
    slug_collisions = find_slug_collisions(candidates, live_items)
    for candidate in candidates:
        candidate["liveSupport"] = live_support_for(candidate["normalized"]["contentType"], catalog)
        candidate["validation"]["issues"] = validate_candidate(candidate)
        candidate["validation"]["status"] = "issues" if candidate["validation"]["issues"] else "ok"
        candidate["duplicateMatches"] = workbook_dupes.get(candidate["dryRunId"], []) + detect_live_duplicates(candidate, catalog)
        if any(collision["dryRunId"] == candidate["dryRunId"] for collision in slug_collisions):
            candidate["manualReviewReasons"].append("slug_collision")
        candidate["disposition"] = classify_disposition(candidate)

    relationships = resolve_relationships(relationships, candidates)
    structural_issues = [{"severity": "warning", "message": message} for message in candidate_warnings + relationship_warnings]
    after_identity = workbook_identity(workbook_path)
    if before_identity["sha256"] != after_identity["sha256"]:
        raise RuntimeError("Workbook hash changed during read-only processing.")

    manifest = {
        "schemaVersion": "1.0",
        "mode": "dry-run",
        "generatedAt": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "workbook": before_identity,
        "workbookAfterProcessing": after_identity,
        "database": catalog.get("database", {}),
        "taxonomy": {
            "mainCategories": MAIN_CATEGORIES,
            "aliases": [{"alias": key, "category": value} for key, value in sorted(ALIAS_TO_CATEGORY.items())],
            "liveSubcategorySupport": catalog.get("support", {}).get("subcategory") == "live_supported",
        },
        "runtimeSupport": catalog.get("support", {}),
        "summary": summarize(candidates, relationships, structural_issues),
        "items": sorted(candidates, key=lambda item: (item["normalized"]["contentType"], item["provenance"]["sourceSheet"], item["provenance"]["excelRow"])),
        "relationships": sorted(relationships, key=lambda item: (item["sourceSheet"], item["excelRow"])),
        "issues": structural_issues,
    }

    write_reports(output_dir, manifest, sheet_inventory, category_plan)
    return manifest


def write_reports(output_dir, manifest, sheet_inventory, category_plan):
    output_dir.mkdir(parents=True, exist_ok=True)
    write_json_safe(output_dir / "content-import-manifest.json", manifest)
    write_json_safe(output_dir / "source-provenance.json", {
        "workbook": manifest["workbook"],
        "sheets": sheet_inventory,
        "contentRows": [
            {
                "dryRunId": item["dryRunId"],
                "sourceSheet": item["provenance"]["sourceSheet"],
                "excelRow": item["provenance"]["excelRow"],
                "workbookContentId": item["provenance"]["workbookContentId"],
            }
            for item in manifest["items"]
        ],
        "categoryPlanRowsSample": category_plan[:60],
    })

    inventory_lines = [
        "# Workbook Inventory",
        "",
        f"Workbook: `{manifest['workbook']['fileName']}`",
        f"SHA-256: `{manifest['workbook']['sha256']}`",
        "",
        "## Sheets",
    ]
    for sheet in sheet_inventory:
        inventory_lines.extend([
            f"- `{sheet['name']}` ({sheet['visibility']}): {sheet['dimensions']['rows']} rows x {sheet['dimensions']['columns']} columns; non-empty rows {sheet['dimensions']['nonEmptyRows']}.",
        ])
        for header in sheet["headerRows"]:
            inventory_lines.append(f"  - Header row {header['row']}: {', '.join(header['headers'])}")
        if sheet["blankRowsInsideDataRanges"]:
            inventory_lines.append(f"  - Blank rows after data begins: {sheet['blankRowsInsideDataRanges']}")
    ignored = [
        "01_Learner_Level_Policy",
        "04_AI_Safety_Test_Set",
        "05_MY_Response_Guidance",
        "06_AI_Quality_Rubric",
        "07_Agent_Tool_Catalogue",
    ]
    inventory_lines.extend(["", "## Import Scope", "Imported into dry-run manifest: `02_Content_Inventory`, `Cyberly_Content_Category_Plan`, `03_Content_Relationships`.", "", "Ignored for import in this phase:"])
    inventory_lines.extend([f"- `{name}`" for name in ignored])
    inventory_lines.extend(["", "## Structural Warnings"])
    if manifest["issues"]:
        inventory_lines.extend([f"- {issue['message']}" for issue in manifest["issues"]])
    else:
        inventory_lines.append("- No blocking workbook structural warnings were detected.")
    (output_dir / "workbook-inventory.md").write_text("\n".join(inventory_lines) + "\n", encoding="utf-8")

    write_csv(output_dir / "taxonomy-mapping.csv", build_taxonomy_rows(manifest["items"]), [
        "original_topic",
        "normalized_topic",
        "mapped_category",
        "live_category_code",
        "proposed_subcategory",
        "mapping_method",
        "confidence",
        "manual_review_note",
    ])

    candidate_rows = []
    validation_rows = []
    duplicate_rows = []
    for item in manifest["items"]:
        duplicate_status = "none"
        if any(match["matchType"] == "exact" for match in item["duplicateMatches"]):
            duplicate_status = "exact"
        elif item["duplicateMatches"]:
            duplicate_status = "possible"
        candidate_rows.append({
            "dry_run_id": item["dryRunId"],
            "sheet": item["provenance"]["sourceSheet"],
            "row": item["provenance"]["excelRow"],
            "content_type": item["normalized"]["contentType"],
            "title": item["normalized"]["title"],
            "original_topic": item["taxonomyMapping"]["original_topic"],
            "mapped_category": item["taxonomyMapping"]["mapped_category"],
            "proposed_subcategory": item["taxonomyMapping"]["proposed_subcategory"],
            "proposed_slug": item["proposedSlug"],
            "live_support": item["liveSupport"],
            "disposition": item["disposition"],
            "validation_status": item["validation"]["status"],
            "duplicate_status": duplicate_status,
            "manual_review_reason": "; ".join(item["manualReviewReasons"]),
        })
        for issue in item["validation"]["issues"]:
            validation_rows.append({
                "severity": issue["severity"],
                "dry_run_id": item["dryRunId"],
                "field": issue["field"],
                "issue_code": issue["issue_code"],
                "message": issue["message"],
                "recommended_action": issue["recommended_action"],
            })
        for match in item["duplicateMatches"]:
            duplicate_rows.append(match)

    write_csv(output_dir / "content-candidates.csv", candidate_rows, [
        "dry_run_id", "sheet", "row", "content_type", "title", "original_topic",
        "mapped_category", "proposed_subcategory", "proposed_slug", "live_support",
        "disposition", "validation_status", "duplicate_status", "manual_review_reason",
    ])
    write_csv(output_dir / "duplicate-report.csv", duplicate_rows, [
        "matchType", "workbookRecordId", "matchedRecordId", "matchedTitle", "reason",
        "fieldsCompared", "confidence", "recommendedDisposition",
    ])
    write_csv(output_dir / "validation-issues.csv", validation_rows, [
        "severity", "dry_run_id", "field", "issue_code", "message", "recommended_action",
    ])

    relationship_rows = []
    for relationship in manifest["relationships"]:
        relationship_rows.append({
            "dry_run_relationship_id": relationship["dryRunRelationshipId"],
            "source_reference": relationship["sourceReference"],
            "relationship_type": relationship["relationshipType"],
            "target_reference": relationship["targetReference"],
            "source_status": relationship["resolution"]["source"]["status"],
            "source_method": relationship["resolution"]["source"]["method"],
            "source_dry_run_id": relationship["resolution"]["source"]["dryRunId"],
            "target_status": relationship["resolution"]["target"]["status"],
            "target_method": relationship["resolution"]["target"]["method"],
            "target_dry_run_id": relationship["resolution"]["target"]["dryRunId"],
            "disposition": relationship["resolution"]["disposition"],
            "ambiguity_notes": relationship["resolution"]["ambiguityNotes"],
        })
    write_csv(output_dir / "relationship-preview.csv", relationship_rows, [
        "dry_run_relationship_id", "source_reference", "relationship_type", "target_reference",
        "source_status", "source_method", "source_dry_run_id", "target_status",
        "target_method", "target_dry_run_id", "disposition", "ambiguity_notes",
    ])

    summary = manifest["summary"]
    summary_lines = [
        "# Dry-Run Import Summary",
        "",
        "This is a read-only planning output. It does not import content, publish content, enable RAG, or modify the database.",
        "",
        "## Candidate Counts",
        f"- Total candidates: {summary['contentCandidates']}",
        f"- By content type: `{json.dumps(summary['byContentType'], sort_keys=True)}`",
        f"- By disposition: `{json.dumps(summary['byDisposition'], sort_keys=True)}`",
        "",
        "## Import Recommendation",
        f"- Future create candidates: {summary['byDisposition'].get('create_candidate', 0)}",
        f"- Future update candidates: {summary['byDisposition'].get('update_candidate', 0)}",
        f"- Exact duplicates to skip: {summary['byDisposition'].get('skip_exact_duplicate', 0)}",
        f"- Manual mapping required: {summary['byDisposition'].get('needs_manual_mapping', 0)}",
        f"- Schema/Admin module required: {summary['byDisposition'].get('unsupported_live_type', 0)}",
        "",
        "## What Must Not Be Imported Yet",
        "- FAQ and Safety Summary rows require live schema/Admin module support before import.",
        "- Scenario rows require detailed step/option modelling before import.",
        "- Workbook governance flags are planning metadata only; every proposed import default remains Draft/Draft/RAG-disabled.",
    ]
    (output_dir / "dry-run-summary.md").write_text("\n".join(summary_lines) + "\n", encoding="utf-8")


def main(argv=None):
    parser = argparse.ArgumentParser(description="Generate Cyberly workbook import preview artifacts without writing to the database.")
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--database-catalog", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args(argv)
    manifest = process(args.workbook, args.database_catalog, args.output)
    print(json.dumps({
        "output": args.output,
        "contentCandidates": manifest["summary"]["contentCandidates"],
        "byDisposition": manifest["summary"]["byDisposition"],
        "workbookHash": manifest["workbook"]["sha256"],
        "databaseUnchanged": manifest["database"].get("unchanged"),
    }, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
